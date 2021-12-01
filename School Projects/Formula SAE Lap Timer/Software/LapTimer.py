from tkinter import ttk, Tk, Label, Button, StringVar, messagebox, filedialog, IntVar, font
import tkinter as tk
import serial
import serial.tools.list_ports
import time
from datetime import datetime, timedelta
from datetime import date
import pandas as pd
import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, colors, Color, PatternFill
import sys
import os.path
from os import path
import threading
import logging
import copy


accelMode = False

accelPassStarted=False
autocrossStarted=False

starttime=None



def set_text(e, text):
    e.delete(0,tk.END)
    e.insert(0,text)
    return

expectedLoss=False
connected=False    
    
def reset():
  global connected
  global newData
  global saved
  global lapTimes
  global timestamps
  global minNum
  global maxNum
  global s
  global dateToday
  
  
  newData=False
  saved=True


  lapTimes = []
  timestamps = []
  
  minNum = None
  maxNum = None



  
  s2=date.today().strftime("%m-%d-%Y")
  s=s2
  dateToday = date.today().strftime("%m-%d-%Y")


  counter=2
  while path.exists("Results/"+s+".xlsx"):
    s = s2 +"-"+str(counter)
    counter += 1


reset()



logFileName="Logs/"+dateToday+".log"



COMport=""
    
def as_text(value):
  if value is None:
    return ""
  return str(value)



class ScrollableFrame(tk.Frame):  
  def __init__(self, container, *args, **kwargs):    
    super().__init__(container, *args, **kwargs)
    self.canvas = tk.Canvas(self, highlightbackground="black", highlightthickness=1)    
    
    def _bound_to_mousewheel(event):
      self.canvas.bind_all("<MouseWheel>", _on_mousewheel)   
    def _unbound_to_mousewheel(event):
      self.canvas.unbind_all("<MouseWheel>")
    def _on_mousewheel(event):    
      self.canvas.yview_scroll(-1*(event.delta//120), "units")        
    scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
    self.scrollable_frame = tk.Frame(self.canvas)
    self.scrollable_frame.bind(
        "<Configure>",
        lambda e: self.canvas.configure(
            scrollregion=self.canvas.bbox("all")
        )
    )
    
    
    
    self.scrollable_frame.pack_propagate(0)
    
    self.canvas.bind('<Enter>', _bound_to_mousewheel)
    self.canvas.bind('<Leave>', _unbound_to_mousewheel)
    
    self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
    self.canvas.configure(yscrollcommand=scrollbar.set)
    self.canvas.pack(side="left", fill="both", expand=True, padx=10)
    scrollbar.pack(side="right", fill="y")    
    
    
def cutHeight(ob):
  height = ob.winfo_height()//2
  width = ob.winfo_width()
  if height < 10 or width < 200:
    height = 10
  elif width < 500 and height > 20:
    height = 15
  elif width < 700 and height > 30:
    height = 30
  else:
    height = 40
  return height
    
class LapTimer: 
  def resetLap(self):  
    global accelPassStarted
    global autocrossStarted
    accelPassStarted=False
    autocrossStarted=False
    self.runningText.grid_forget()
    self.reset_lap_button.grid_forget()
  def cb(self):   
    global accelMode
    global accelPassStarted
    global autocrossStarted
    
    accelPassStarted=False
    autocrossStarted=False
    self.runningText.grid_forget()
    self.reset_lap_button.grid_forget()
    
    if self.modeVar.get()==1:
      #tell receiver to use accel mode      
      accelMode=True
    else:      
      #tell receiver to use autocross mode      
      accelMode=False
  def resizeFont(self, event): 
    try:
      self.fontConBut['size'] = cutHeight(self.connect_button)
      self.fontRunning['size'] = cutHeight(self.runningText)
      self.fontResetBut['size'] = cutHeight(self.reset_lap_button)
      self.fontConText['size'] = cutHeight(self.connectionLbl)
      self.fontDate['size'] = cutHeight(self.dateLbl)
      self.fontLoc['size'] = cutHeight(self.locationLbl)
      self.fontEvent['size'] = cutHeight(self.eventLbl)
      self.fontNotesLbl['size'] = cutHeight(self.notesLbl)
      self.fontTitle['size'] = cutHeight(self.label)
      
      self.scroll.scrollable_frame.config(width=self.scroll.canvas.winfo_width()-1)
      
    except:
      pass    
  def resend1(self):
    ser.write(b'1')
  def resend2(self):
    ser.write(b'2')
  def __init__(self, master):
    global COMport
    self.master = master
    self.master.bind('<Configure>', self.resizeFont)
    
    self.rowOne = tk.Frame(master)
    self.rowOne.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.rowOne.grid_propagate(0)
    self.rowTwo = tk.Frame(master,bg="green")
    self.rowTwo.grid(row=1, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.rowTwo.grid_propagate(0)
    self.rowThree = tk.Frame(master,bg="blue")
    self.rowThree.grid(row=2, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.rowThree.grid_propagate(0)
    master.rowconfigure(0, weight=100)
    master.rowconfigure(1, weight=629)
    master.rowconfigure(2, weight=234)
    master.columnconfigure(0, weight=1)
    
    self.conSignFrame = tk.Frame(self.rowOne)
    self.conSignFrame.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.conSignFrame.grid_propagate(0)
    self.titleFrame = tk.Frame(self.rowOne)
    self.titleFrame.grid(row=0, column=1, sticky=tk.E+tk.W+tk.N+tk.S);
    self.titleFrame.grid_propagate(0)
    self.progressFrame = tk.Frame(self.rowOne)
    self.progressFrame.grid(row=0, column=2, sticky=tk.E+tk.W+tk.N+tk.S);
    self.progressFrame.grid_propagate(0)
    self.resetFrame = tk.Frame(self.rowOne)
    self.resetFrame.grid(row=0, column=3, sticky=tk.E+tk.W+tk.N+tk.S);    
    self.resetFrame.grid_propagate(0)
    self.rowOne.columnconfigure(0, weight=550)
    self.rowOne.columnconfigure(1, weight=485)
    self.rowOne.columnconfigure(2, weight=385)
    self.rowOne.columnconfigure(3, weight=180)
    self.rowOne.rowconfigure(0, weight=1)
    
    self.infoFrame = tk.Frame(self.rowTwo)
    self.infoFrame.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.infoFrame.grid_propagate(0)
    self.notesFrame = tk.Frame(self.rowTwo)
    self.notesFrame.grid(row=0, column=1, sticky=tk.E+tk.W+tk.N+tk.S);
    self.notesFrame.grid_propagate(0)
    self.rowTwo.columnconfigure(0, weight=920)
    self.rowTwo.columnconfigure(1, weight=680)
    self.rowTwo.rowconfigure(0, weight=1)
    
    self.connectFrame = tk.Frame(self.rowThree)
    self.connectFrame.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S);
    self.connectFrame.grid_propagate(0)
    self.avgFrame = tk.Frame(self.rowThree)
    self.avgFrame.grid(row=0, column=1, sticky=tk.E+tk.W+tk.N+tk.S);
    self.avgFrame.grid_propagate(0)
    self.clearFrame = tk.Frame(self.rowThree)
    self.clearFrame.grid(row=0, column=2, sticky=tk.E+tk.W+tk.N+tk.S);
    self.clearFrame.grid_propagate(0)
    self.exportFrame = tk.Frame(self.rowThree)
    self.exportFrame.grid(row=0, column=3, sticky=tk.E+tk.W+tk.N+tk.S); 
    self.exportFrame.grid_propagate(0)

    self.rowThree.columnconfigure(0, weight=500)
    self.rowThree.columnconfigure(1, weight=400)
    self.rowThree.columnconfigure(2, weight=380)
    self.rowThree.columnconfigure(3, weight=320)
    self.rowThree.rowconfigure(0, weight=1)
    
    
    
    
    
    master.title("FSAE Lap Timer")
    
    self.fontTitle = font.Font(self.master, family='Arial', size=12)
    self.label = Label(self.titleFrame, text="Formula SAE Lap Timer", font = self.fontTitle)           
    self.label.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S)
    self.titleFrame.columnconfigure(0, weight=1)
    self.titleFrame.rowconfigure(0, weight=1)
    
    self.fontRunning = font.Font(self.master, family='Arial', size=12)
    self.runningText = Label(self.progressFrame, text="Lap in progress", fg="green", font=self.fontRunning)    
    self.progressFrame.columnconfigure(0, weight=1)
    self.progressFrame.rowconfigure(0, weight=1)
    
    self.fontResetBut = font.Font(self.master, family='Arial', size=12)
    self.reset_lap_button = Button(self.resetFrame, text="Reset Lap", command=self.resetLap, font=self.fontResetBut)        
    self.resetFrame.columnconfigure(0, weight=1)
    self.resetFrame.columnconfigure(1, weight=1)
    self.resetFrame.columnconfigure(2, weight=1)
    self.resetFrame.rowconfigure(0, weight=1)
    self.resetFrame.rowconfigure(1, weight=1)
    self.resetFrame.rowconfigure(2, weight=1)
    
    self.fontConText = font.Font(self.master, family='Arial', size=12)
    self.connectionText = StringVar()    
    self.connectionText.set("Receiver disconnected")
    self.connectionLbl = Label(self.conSignFrame, textvariable=self.connectionText, font=self.fontConText, fg="red")           
    self.connectionLbl.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S)
    self.conSignFrame.columnconfigure(0, weight=1)
    self.conSignFrame.rowconfigure(0, weight=1)
    
    
    self.scroll = ScrollableFrame(self.infoFrame)    
    self.graphHead = DataFrame(self.infoFrame, True, padx=10)
    self.graphHead.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S, padx=(0,17))    
    self.scroll.grid(row=1, column=0, sticky=tk.E+tk.W+tk.N+tk.S)    
    self.infoFrame.columnconfigure(0, weight=1)
    self.infoFrame.rowconfigure(0, weight=100)
    self.infoFrame.rowconfigure(1, weight=550)    
    
    #DataFrame(self.scroll.canvas,bg="green").pack(fill="x")
    
    
    
    
    self.fontDate = font.Font(self.master, family='Arial', size=12)
    self.dateLbl = Label(self.notesFrame, text="Date: ", font= self.fontDate)           
    self.dateLbl.grid(row=0, column=0, sticky=tk.E+tk.S, pady=10)      
    self.fontLoc = font.Font(self.master, family='Arial', size=12)
    self.locationLbl = Label(self.notesFrame, text="Location: ", font=self.fontLoc)           
    self.locationLbl.grid(row=1, column=0, sticky=tk.E+tk.N+tk.S)
    self.fontEvent = font.Font(self.master, family='Arial', size=12)
    self.eventLbl = Label(self.notesFrame, text="Event: ", font=self.fontEvent)           
    self.eventLbl.grid(row=2, column=0, sticky=tk.E+tk.N+tk.S)   
    self.fontNotesLbl = font.Font(self.master, family='Arial', size=12)
    self.notesLbl = Label(self.notesFrame, text="Notes: ", font= self.fontNotesLbl)           
    self.notesLbl.grid(row=3, column=0, sticky=tk.E+tk.N, pady=10)    
    
    self.notesFrame.columnconfigure(0, weight=176)
    self.notesFrame.columnconfigure(1, weight=504)
    self.notesFrame.rowconfigure(0, weight=80)
    self.notesFrame.rowconfigure(1, weight=63)
    self.notesFrame.rowconfigure(2, weight=63)
    self.notesFrame.rowconfigure(3, weight=430)
    
    
    self.date = tk.Entry(self.notesFrame)
    self.date.grid(row=0, column=1, sticky=tk.E+tk.S+tk.W, pady=11, padx=20)
    
    self.location = tk.Entry(self.notesFrame)
    self.location.grid(row=1, column=1, sticky=tk.E+tk.W, padx=20)
    
    self.event = tk.Entry(self.notesFrame)
    self.event.grid(row=2, column=1, sticky=tk.E+tk.W, padx=20)
    
    self.notesInsideFrame = tk.Frame(self.notesFrame)
    self.notesInsideFrame.grid(row=3, column=1, sticky=tk.N+tk.E+tk.S+tk.W)     
    self.notes = tk.Text(self.notesInsideFrame, font=("Arial", 8))
    self.notes.pack(side=tk.LEFT, fill=tk.BOTH, expand=tk.YES, padx=20, pady=10) 
    self.notesInsideFrame.pack_propagate(0)
    set_text(self.date, dateToday)
    
    
    
    
    self.btn_text = StringVar()
    self.btn_text.set("Connect to receiver")    
    self.fontConBut = font.Font(self.master, family='Arial', size=12)
    self.connect_button = Button(self.connectFrame, textvar=self.btn_text, command=tryConnect, font=self.fontConBut)    
    self.connect_button.grid(row=1, column=1, sticky=tk.E+tk.W+tk.N+tk.S)        
    self.connectFrame.grid_propagate(0)
    self.connectFrame.columnconfigure(0, weight=1)
    self.connectFrame.columnconfigure(1, weight=1)
    self.connectFrame.columnconfigure(2, weight=1)
    self.connectFrame.rowconfigure(0, weight=1)
    self.connectFrame.rowconfigure(1, weight=1)
    self.connectFrame.rowconfigure(2, weight=1)
    
    
    self.avgText = StringVar()    
    self.avgText.set("Average time: 0")
    self.avg = Label(self.avgFrame, textvariable=self.avgText, font=("Arial", 20))           
    self.avg.grid(row=0, column=0, columnspan=6, sticky=tk.E+tk.W+tk.N+tk.S)       
    self.modeVar = IntVar()
    self.mode = tk.Checkbutton(self.avgFrame, text="Accel test mode", variable=self.modeVar,font=("Arial", 20), command=self.cb)
    self.mode.grid(row=1, column=0, columnspan=6, sticky=tk.E+tk.W+tk.N+tk.S)  
    
    self.resend1 = Button(self.avgFrame, text="Resend 1", command=self.resend1,font=("Arial", 16))
    self.resend1.grid(row=3, column=1, sticky=tk.E+tk.W+tk.N+tk.S)  
    
    self.resend2 = Button(self.avgFrame, text="Resend 2", command=self.resend2,font=("Arial", 16))
    self.resend2.grid(row=3, column=4, sticky=tk.E+tk.W+tk.N+tk.S)  
    
    self.avgFrame.columnconfigure(0, weight=1)
    self.avgFrame.columnconfigure(1, weight=1)
    self.avgFrame.columnconfigure(2, weight=1)
    self.avgFrame.columnconfigure(3, weight=1)
    self.avgFrame.columnconfigure(4, weight=1)
    self.avgFrame.columnconfigure(5, weight=1)
    self.avgFrame.rowconfigure(0, weight=1)
    self.avgFrame.rowconfigure(1, weight=1)
    self.avgFrame.rowconfigure(2, weight=1)
    self.avgFrame.rowconfigure(3, weight=1)
    self.avgFrame.rowconfigure(4, weight=1)
    
    
    
    
    self.reset_button = Button(self.clearFrame, text="Clear Data", command=rst,font=("Arial", 16))
    self.reset_button.grid(row=1, column=1, sticky=tk.E+tk.W+tk.N+tk.S)  
    self.clearFrame.columnconfigure(0, weight=1)
    self.clearFrame.columnconfigure(1, weight=1)
    self.clearFrame.columnconfigure(2, weight=1)
    self.clearFrame.rowconfigure(0, weight=1)
    self.clearFrame.rowconfigure(1, weight=1)
    self.clearFrame.rowconfigure(2, weight=1)
    
    self.save_button = Button(self.exportFrame, text="Save Excel Sheet", command=save, font=("Arial", 16))
    self.save_button.grid(row=1, column=1, sticky=tk.E+tk.W+tk.N+tk.S)  
    self.exportFrame.columnconfigure(0, weight=1)
    self.exportFrame.columnconfigure(1, weight=1)
    self.exportFrame.columnconfigure(2, weight=1)
    self.exportFrame.rowconfigure(0, weight=1)
    self.exportFrame.rowconfigure(1, weight=1)
    self.exportFrame.rowconfigure(2, weight=1)
    
        
        
class DataFrame(tk.Frame):
  header=False
  def dele(self, event):
    global lapTimes
    global timestamps
    global maxNum
    global minNum
    if messagebox.askokcancel("Delete", "Delete Lap?"):
    
      v=int(self.lapNumText.get()[4:])-1
      
      lapTimes = lapTimes[0:v]+lapTimes[v+1:len(lapTimes)]
      timestamps = timestamps[0:v]+timestamps[v+1:len(timestamps)]
      
      
      self.destroy()
      avg=0
      for t in lapTimes:
        avg+=t
      
      if len(lapTimes)>0:
        avg /= len(lapTimes)
        minNum = min(lapTimes)
        maxNum = max(lapTimes)
      else:
        minNum = None
        maxNum = None
      my_gui.avgText.set("Average time: "+str(round(avg,4)))
      
      i=1      
      for widget in self.master.winfo_children():
        widget.lapNumLbl.config(fg="black")
        if (widget.header==False):
          if float(widget.lapDurText.get()) == maxNum:          
            widget.lapNumLbl.config(fg="red")
          if float(widget.lapDurText.get()) == minNum:
            widget.lapNumLbl.config(fg="green")
          widget.lapNumText.set("Lap "+str(i))          
          i+=1
      my_gui.scroll.scrollable_frame.config(height=40*len(my_gui.scroll.scrollable_frame.winfo_children()))
      
    
    
    
    
  def __init__(self, container, head=False, *args, **kwargs):    
    super().__init__(container, *args, **kwargs)
    self.header=head
    
    
    
    
    
    
    
    self.lapNum = tk.Frame(self, height=40, highlightbackground="black", highlightthickness=1)
    self.lapNum.grid(row=0, column=0, sticky=tk.E+tk.W+tk.N+tk.S) 
    self.lapNumText = StringVar()
    self.lapNumText.set("Lap #")
    self.lapNum.grid_propagate(0)
    self.lapNumLbl = Label(self.lapNum, textvariable=self.lapNumText)
    self.lapNumLbl.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    
    if not self.header:
      self.lapNum.bind("<Button-1>", self.dele)
    if not self.header:
      self.lapNumLbl.bind("<Button-1>", self.dele)
    
    self.time = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)    
    if not self.header:
      self.time.bind("<Button-1>", self.dele)
    self.time.grid(row=0, column=1, sticky=tk.E+tk.W+tk.N+tk.S)    
    self.timeText = StringVar()
    self.timeText.set("Time\nRecorded")
    self.time.grid_propagate(0)
    self.timeLbl = Label(self.time, textvariable=self.timeText)
    self.timeLbl.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    if not self.header:
      self.timeLbl.bind("<Button-1>", self.dele)
    
    
    self.lapDur = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)    
    if not self.header:
      self.lapDur.bind("<Button-1>", self.dele)
    self.lapDur.grid(row=0, column=2, sticky=tk.E+tk.W+tk.N+tk.S)     
    self.lapDurText = StringVar()
    self.lapDurText.set("Lap\nDuration (s)")
    self.lapDur.grid_propagate(0)
    self.lapDurLbl = Label(self.lapDur, textvariable=self.lapDurText)
    self.lapDurLbl.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    if not self.header:
      self.lapDurLbl.bind("<Button-1>", self.dele)
    
    if self.header:
      self.driver = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)    
      self.driver.grid(row=0, column=3, sticky=tk.E+tk.W+tk.N+tk.S)    
      self.driverText = StringVar()
      self.driverText.set("Driver")
      self.driver.grid_propagate(0)
      Label(self.driver, textvariable=self.driverText).place(relx=0.5, rely=0.5, anchor=tk.CENTER)
      
      self.notes = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)    
      self.notes.grid(row=0, column=4, sticky=tk.E+tk.W+tk.N+tk.S)    
      self.notesText = StringVar()
      self.notesText.set("Notes")
      self.notes.grid_propagate(0)
      Label(self.notes, textvariable=self.notesText).place(relx=0.5, rely=0.5, anchor=tk.CENTER)
      
    else:
      self.driver = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)    
      self.driver.grid(row=0, column=3, sticky=tk.E+tk.W+tk.N+tk.S) 
      self.driver.grid_propagate(0)
      self.driverText = tk.Entry(self.driver) 
      self.driverText.grid(row=0, column=0, sticky=tk.W+tk.E, padx=5)
      self.driver.columnconfigure(0, weight=1)
      self.driver.rowconfigure(0, weight=1)
      
      
      self.notes = tk.Frame(self,height=40, highlightbackground="black", highlightthickness=1)          
      self.notes.grid(row=0, column=4, sticky=tk.E+tk.W+tk.N+tk.S)    
      self.notes.grid_propagate(0)
      self.notesText = tk.Text(self.notes, height=3, width=32, font=("Arial", 8))
      self.notesText.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S, padx=5, pady=5)
      self.notes.columnconfigure(0, weight=1)
      self.notes.rowconfigure(0, weight=1)


    self.columnconfigure(0, weight=125)
    self.columnconfigure(1, weight=125)
    self.columnconfigure(2, weight=125)
    self.columnconfigure(3, weight=250)
    self.columnconfigure(4, weight=250)
    self.rowconfigure(0, weight=1)

    
      
    
    
    
    





    
        
        
        
        
root = Tk()




def on_closing():
  if saved:    
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
      now = datetime.now()
      current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
      f = open(logFileName, "a")      
      f.write(current_time+": "+"Receiver disconnected"+"\n")
      f.close()
      root.destroy()
  else:
    if messagebox.askokcancel("Quit", "Are you sure you want to quit without saving?"):
      now = datetime.now()
      current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
      f = open(logFileName, "a")      
      f.write(current_time+": "+"Receiver disconnected"+"\n")
      f.close()
      root.destroy()
    
root.protocol("WM_DELETE_WINDOW", on_closing)
#root.geometry("1280x768")
root.geometry("1024x768")
#root.resizable(False, False)

startOfConnection = datetime.now()


def connect():
  global ser
  global connected
  global tryingToConnect
  global logFileName
  global startOfConnection
  ports = serial.tools.list_ports.comports()
  found=False  
  for port, desc, hwid in sorted(ports):
    try:      
      my_gui.connectionText.set("Trying "+port)
      my_gui.connectionLbl.config(fg="#babd00")
      ser = serial.Serial(port=port, baudrate=9600, timeout=3)
      ser_bytes = ser.readline()      
      decoded_bytes = ser_bytes[0:len(ser_bytes)-2]
      msg = decoded_bytes.decode("utf-8")
      if msg == "SAE Connected":        
        now = datetime.now()
        startOfConnection = now
        
        
        current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
        f = open(logFileName, "a")      
        f.write(current_time+": "+"Receiver connected"+"\n")
        f.close()
        found=True
        COMport = port
        break
    except UnicodeDecodeError:
      print("Got response for " + str(port) + ", but message was mangled")
    except:
      pass
  if not found:    
    my_gui.connectionText.set("Receiver disconnected")
    my_gui.connectionLbl.config(fg="red")
    tryingToConnect=False
    messagebox.showinfo("No Connection", "The receiver was not detected")
    exit(0)
  connected=True
  my_gui.connectionText.set("Connected: "+COMport)
  my_gui.connectionLbl.config(fg="green")

  x = threading.Thread(target=listen, daemon=True)
  x.start()
  root.after(10, task)
  my_gui.btn_text.set("Disconnect from receiver")
  tryingToConnect=False

tryingToConnect=False
def tryConnect():
  global connected
  global tryingToConnect
  global ser
  global expectedLoss
  if tryingToConnect == False and connected == False:
    try:
      tryingToConnect=True
      x = threading.Thread(target=connect, daemon=True)
      x.start()
    except:
      tryingToConnect=False
      pass    
  elif tryingToConnect == False and connected:    
    #disconnect
    expectedLoss=True
    connected=False    
    ser.close()
  
def rst():
  global accelPassStarted
  global autocrossStarted

  if messagebox.askokcancel("Reset", "Are you sure you want to clear this data?"):
    reset()
    accelPassStarted=False
    autocrossStarted=False
    for widget in my_gui.scroll.scrollable_frame.winfo_children():
      if widget.header == False:
        widget.destroy()
    my_gui.avgText.set("Average time: 0")    
    set_text(my_gui.date, dateToday)
    set_text(my_gui.location, "")
    set_text(my_gui.event, "")        
    my_gui.notes.delete(1.0,"end")  
    my_gui.runningText.grid_forget()
    my_gui.reset_lap_button.grid_forget()
    

def save():
  global saved
  f = filedialog.asksaveasfilename(defaultextension=".xlsx",  filetypes=(("excel file", "*.xlsx"),), initialfile=s+".xlsx")  
  if f!="":
    writeExcel(f)
    saved=True

my_gui = LapTimer(root)

def writeExcel(f):
  global dateToday
  global location
  global event
  global mainNotes
  
  l = len(lapTimes)+1

  driv=[]
  notes=[]
  c = my_gui.scroll.scrollable_frame.winfo_children()
  for i in range(l-1):
    driv = driv + [c[i].driverText.get()]
    notes = notes + [c[i].notesText.get("1.0", tk.END)]
  driv = driv + [""]
  notes = notes + [""]
  
  
  d = {'Time Lap Recorded': timestamps + ["Average Lap Time:"], 'Lap Duration (s)': lapTimes+["=AVERAGE(C2:C"+str(l)+")"], 'Difference from Average (s)' : ["=C"+str(i+2)+"-C"+str(l+1) for i in range(l-1)]+[""], "Driver" : driv,"Notes" : notes}
  df = pd.DataFrame(data=d)
  df.index += 1
           
  
  writer_args = {
        'path': f,
        'mode': 'w',
        'engine': 'openpyxl'}
    
  with pd.ExcelWriter(**writer_args) as xlsx:
    df.to_excel(xlsx, s)
    
    ws = xlsx.sheets[s]
    
    
    
    for row in range(3):
      ws.merge_cells(start_row=row+1, start_column=8, end_row=row+1, end_column=11)
      ws['H'+str(row+1)].font = Font(bold=True)      
    ws.merge_cells(start_row=4, start_column=8, end_row=20, end_column=11)    
    ws['H1']="Date: "+my_gui.date.get()
    ws['H2']="Location: "+my_gui.location.get()
    ws['H3']="Event: "+my_gui.event.get()
    ws['H4']="Notes: "+my_gui.notes.get("1.0", tk.END)
    ws['H4'].alignment = Alignment(horizontal='left', vertical='top')
    
    
    ws['A1'] = "Lap #"
    ws['A1'].alignment = Alignment(horizontal='center')
    
    
    for c in range(5):
      for r in range(l):      
        ws.cell(row=2+r, column=2+c).alignment = Alignment(horizontal='right')
        
    
    ws['A1'].font = Font(bold=True)
    ws["A"+str(l+1)] = ""
    ws["A"+str(l+1)].border = Border(top=Side(style='thin'))
    
    redFill = PatternFill(start_color='FFC00000',
                   end_color='FFFF0000',
                   fill_type='solid')
    greenFill = PatternFill(start_color='FF00B050',
                   end_color='FF00B050',
                   fill_type='solid')
    
    
    for row in range(len(lapTimes)):
      if lapTimes[row] == maxNum:
        ws.cell(row=2+row, column=1).fill = redFill
    
    for row in range(len(lapTimes)):
      if lapTimes[row] == minNum:
        ws.cell(row=2+row, column=1).fill = greenFill
    
    
    for col in ['A','B','C','D']:
      length = max(len(as_text(cell.value)) for cell in ws[col])      
      ws.column_dimensions[col].width = length 
    
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30
    
    
    a = copy.copy(ws.cell(row=4, column=8).alignment)
    a.wrapText=True
    ws.cell(row=4, column=8).alignment = a






def task():
  global newData
  global lapTimes
  global timestamps
  global connected
  global saved
  global expectedLoss
  if newData:
    newData=False
    saved=False
    f=DataFrame(my_gui.scroll.scrollable_frame)       
    f.lapNumText.set("Lap "+str(len(lapTimes)))      
    f.timeText.set(timestamps[len(lapTimes)-1])
    f.lapDurText.set("{:.3f}".format(lapTimes[len(lapTimes)-1]))
    f.pack(fill="x")
    my_gui.scroll.scrollable_frame.config(height=40*len(my_gui.scroll.scrollable_frame.winfo_children()))
    for widget in my_gui.scroll.scrollable_frame.winfo_children():
      widget.lapNumLbl.config(fg="black")
      if widget.header == False:
        if float(widget.lapDurText.get()) == maxNum:          
          widget.lapNumLbl.config(fg="red")
        if float(widget.lapDurText.get()) == minNum:
          widget.lapNumLbl.config(fg="green")
    
    avg=0
    for v in lapTimes:
      avg += v
    avg /= len(lapTimes)
    my_gui.avgText.set("Average time: "+str(round(avg,4)))
    
  if not connected:
    my_gui.btn_text.set("Connect to receiver")
    my_gui.connectionText.set("Receiver disconnected")
    my_gui.connectionLbl.config(fg="red")
    if expectedLoss:
      expectedLoss=False
      messagebox.showerror("Connection Terminated", "The connection with the receiver was severed")
      now = datetime.now()
      current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
      f = open(logFileName, "a")      
      f.write(current_time+": "+"Receiver disconnected"+"\n")
      f.close()
    else:
      messagebox.showerror("Connection Lost", "The connection with the receiver was lost")
      now = datetime.now()
      current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
      f = open(logFileName, "a")      
      f.write(current_time+": "+"Receiver disconnected"+"\n")
      f.close()
    return
    
    
    
  
  
  
  root.after(10, task)




def listen():
  global newData
  global minNum
  global maxNum
  global timestamps
  global lapTimes
  global connected
  global ser
  global starttime
  global accelMode
  global accelPassStarted
  global autocrossStarted
    
  ser.timeout = 1

   

  
  while True:
    try:
      ser_bytes = ser.readline()      
      decoded_bytes = ser_bytes[0:len(ser_bytes)-2]
      msg = decoded_bytes.decode("utf-8")          
      
      now = datetime.now()
      current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
      if (msg != ""):
        f = open(logFileName, "a")      
        f.write(current_time+": "+msg+"\n")
        f.close()
      
      if msg[0:3] == "T: ":                
        
        val = float(msg[3:msg.find(",")])        
        t = now - timedelta(milliseconds=val)                
        fromNum = int(msg[msg.find("From ")+5:]) 
        
        
        def addLap(l):
          global newData
          global minNum
          global maxNum
          global timestamps
          global lapTimes
          lapTimes = lapTimes + [l]                  
          current_time = now.strftime("%I:%M:%S.%f")+now.strftime("%p").lower()
          timestamps = timestamps + [current_time]                
          if minNum==None:
            minNum = l
            maxNum = l
          else:
            if l < minNum:
              minNum = l
            if l > maxNum:
              maxNum = l               
          newData=True 
        
        if not accelMode:
          if not autocrossStarted:
            my_gui.runningText.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
            my_gui.reset_lap_button.grid(row=1, column=1, sticky=tk.E+tk.W+tk.N+tk.S)
            autocrossStarted=True
            starttime = t
          else:
            addLap(round(abs((t-starttime).total_seconds()), 3))
            starttime = t          
        else:
          if not accelPassStarted:
            accelPassStarted=True
            my_gui.runningText.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
            my_gui.reset_lap_button.grid(row=1, column=1, sticky=tk.E+tk.W+tk.N+tk.S)
            starttime = t
          else:
            addLap(round(abs((t-starttime).total_seconds()),3))
            my_gui.runningText.grid_forget()
            my_gui.reset_lap_button.grid_forget()
            accelPassStarted=False         
        
        
        
    except serial.serialutil.SerialException:      
      connected=False
      
      exit(0)      
    except:
      connected=False      
      if not expectedLoss:
        logging.error("Unrecoverable error occurred in the serial thread")   
      exit(0)
   
logging.basicConfig(level=logging.DEBUG)    


  
  
root.after(5, tryConnect)
root.mainloop()


















